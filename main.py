# range_paster_gui_professional.py
# Requires: pip install openpyxl
import os
import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from typing import Tuple, Optional, List, Dict, Any
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ----------------------------- Module logger -----------------------------
LOGGER = None  # set to a callable(str) by the GUI so helpers can log to the Output Log


def set_logger(fn):
    global LOGGER
    LOGGER = fn


def log_debug(msg: str):
    if LOGGER:
        LOGGER(msg)
    else:
        print(msg)


# ----------------------------- Core helpers -----------------------------

def last_nonempty_col_in_row_range(ws: Worksheet, start_col: int, min_row: int, max_row: int) -> int:
    """
    Scan columns from start_col through ws.max_column, and within the given row range [min_row..max_row],
    return the last column index that contains any non-empty cell.
    """
    last = start_col
    max_scan_col = ws.max_column
    for c in range(start_col, max_scan_col + 1):
        for r in range(min_row, max_row + 1):
            if cell_is_nonempty(ws, r, c):
                last = c
                break
    log_debug(
        f"[calc] last_nonempty_col_in_row_range rows={min_row}..{max_row} start_col={start_col} -> last_col={last}")
    return last


def format_range_a1(min_c: int, min_r: int, max_c: int, max_r: int, sheet: Optional[str]) -> str:
    start = f"{get_column_letter(min_c)}{min_r}"
    end = f"{get_column_letter(max_c)}{max_r}"
    core = start if (min_c == max_c and min_r == max_r) else f"{start}:{end}"
    return f"{sheet}!{core}" if sheet else core


def parse_sheet_and_ref(token: str) -> Tuple[Optional[str], str]:
    token = token.strip()
    if "!" in token:
        sheet, ref = token.split("!", 1)
        res = (sheet.strip(), ref.strip())
        log_debug(f"[calc] parse_sheet_and_ref '{token}' -> sheet='{res[0]}', ref='{res[1]}'")
        return res
    log_debug(f"[calc] parse_sheet_and_ref '{token}' -> active sheet, ref='{token}'")
    return None, token


def parse_cell(ref: str) -> Tuple[int, int]:
    from openpyxl.utils.cell import coordinate_to_tuple
    row, col = coordinate_to_tuple(ref)
    log_debug(f"[calc] parse_cell '{ref}' -> col={col}, row={row}")
    return col, row


def parse_range_standard(ref: str) -> Tuple[int, int, int, int]:
    from openpyxl.utils.cell import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    log_debug(f"[calc] parse_range_standard '{ref}' -> ({min_col},{min_row})..({max_col},{max_row})")
    return min_col, min_row, max_col, max_row


def get_sheet(wb, name: Optional[str]) -> Worksheet:
    if name is None:
        ws = wb.active
        log_debug(f"[calc] get_sheet active -> '{ws.title}'")
        return ws
    if name in wb.sheetnames:
        log_debug(f"[calc] get_sheet found -> '{name}'")
        return wb[name]
    msg = f'Sheet "{name}" not found in workbook.'
    log_debug(f"[calc] get_sheet error -> {msg}")
    raise ValueError(msg)


def cell_is_nonempty(ws: Worksheet, row: int, col: int) -> bool:
    v = ws.cell(row=row, column=col).value
    return v is not None and (not isinstance(v, str) or v.strip() != "")


def last_nonempty_col_on_row(ws: Worksheet, row: int, start_col: int) -> int:
    """
    Scan columns on a single row from start_col through ws.max_column and return the last non-empty column index.
    """
    last = start_col
    max_scan_col = ws.max_column
    for c in range(start_col, max_scan_col + 1):
        if cell_is_nonempty(ws, row, c):
            last = c
    log_debug(f"[calc] last_nonempty_col_on_row row={row} from_col={start_col} -> last_col={last}")
    return last


def last_nonempty_row_on_col(ws: Worksheet, col: int, start_row: int) -> int:
    last = start_row
    for r in range(start_row, ws.max_row + 1):
        if cell_is_nonempty(ws, r, col):
            last = r
    log_debug(f"[calc] last_nonempty_row_on_col col={col} from_row={start_row} -> last_row={last}")
    return last


def first_nonempty_row_on_col(ws: Worksheet, col: int, start_row: int = 1) -> int:
    """
    Scan downward from start_row on a single column and return the first row index
    that has a non-empty cell. If none found, return start_row.
    """
    for r in range(start_row, ws.max_row + 1):
        if cell_is_nonempty(ws, r, col):
            log_debug(f"[calc] first_nonempty_row_on_col col={col} -> first_row={r}")
            return r
    log_debug(f"[calc] first_nonempty_row_on_col col={col} -> none found, default={start_row}")
    return start_row


def last_nonempty_row_in_col_range(ws: Worksheet, min_col: int, max_col: int, start_row: int) -> int:
    """
    Within columns [min_col..max_col], find the last row at or below start_row
    where ANY column in that span has a non-empty cell.
    """
    last = start_row
    for r in range(start_row, ws.max_row + 1):
        any_nonempty = False
        for c in range(min_col, max_col + 1):
            if cell_is_nonempty(ws, r, c):
                any_nonempty = True
                break
        if any_nonempty:
            last = r
    log_debug(
        f"[calc] last_nonempty_row_in_col_range cols={min_col}..{max_col} from_row={start_row} -> last_row={last}")
    return last


def split_col_row(token: str) -> Tuple[Optional[str], Optional[str]]:
    token = token.strip()
    if token == "??":
        log_debug(f"[calc] split_col_row '{token}' -> col='?', row='?'")
        return "?", "?"
    i = 0
    while i < len(token) and token[i].isalpha():
        i += 1
    col_part = token[:i] if i > 0 else None
    row_part = token[i:] if i < len(token) else None
    if col_part is None and token and token[0] == "?":
        col_part = "?"
        row_part = token[1:] if len(token) > 1 else None
    if row_part is None and token and token[-1] == "?":
        row_part = "?"
    log_debug(f"[calc] split_col_row '{token}' -> col='{col_part}', row='{row_part}'")
    return col_part, row_part


def parse_range_with_wildcards_basic(ws: Worksheet, ref: str) -> Tuple[int, int, int, int]:
    """
    Original wildcard handler: A23:B? or A23:?56 or A23:??
    Uses row/col expansion based on existing non-empty content.
    """
    left, right = [x.strip() for x in ref.split(":", 1)]
    start_col, start_row = parse_cell(left)
    end_col_tok, end_row_tok = split_col_row(right)

    if end_col_tok == "?":
        end_col = last_nonempty_col_on_row(ws, start_row, start_col)
    elif end_col_tok is None:
        end_col = start_col
    else:
        end_col = column_index_from_string(end_col_tok)

    if end_row_tok == "?":
        end_row = last_nonempty_row_on_col(ws, start_col, start_row)
    elif end_row_tok is None:
        end_row = start_row
    else:
        try:
            end_row = int(end_row_tok)
        except ValueError:
            raise ValueError(f'Invalid row in "{right}"')

    min_c, max_c = sorted((start_col, end_col))
    min_r, max_r = sorted((start_row, end_row))
    log_debug(f"[calc] parse_range_with_wildcards_basic '{ref}' -> ({min_c},{min_r})..({max_c},{max_r})")
    return min_c, min_r, max_c, max_r


def copy_values(src_ws: Worksheet, dst_ws: Worksheet,
                src_bounds: Tuple[int, int, int, int],
                dst_row: int, dst_col: int) -> Tuple[int, int]:
    min_c, min_r, max_c, max_r = src_bounds
    width = max_c - min_c + 1
    height = max_r - min_r + 1  # new comment
    log_debug(
        f"[calc] copy_values bounds=({min_c},{min_r})..({max_c},{max_r}) -> width={width}, height={height}, paste_to=({dst_col},{dst_row})")

    for r_off in range(height):
        for c_off in range(width):
            v = src_ws.cell(row=min_r + r_off, column=min_c + c_off).value
            dst_ws.cell(row=dst_row + r_off, column=dst_col + c_off).value = v

    return width, height


def normalize_value(v: Any) -> str:
    """Stringify and strip to compare values robustly."""
    if v is None:
        return ""
    return str(v).strip()


def find_row_by_value(ws: Worksheet, search_value: Any) -> Optional[int]:
    """
    Search for search_value across columns 1..ws.max_column, but skip columns that are entirely empty.
    Returns the 1-based row index of the first match (top-down), or None if not found.
    """
    target = normalize_value(search_value)
    if not target:
        log_debug(f"[search] find_row_by_value empty target -> None")
        return None

    max_col = ws.max_column
    log_debug(f"[search] find_row_by_value target='{target}' scanning cols 1..{max_col}, rows 1..{ws.max_row}")

    for col in range(1, max_col + 1):
        # Skip columns that are completely empty to avoid wasted scans
        any_nonempty = False
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=col).value is not None:
                any_nonempty = True
                break
        if not any_nonempty:
            continue

        for r in range(1, ws.max_row + 1):
            if normalize_value(ws.cell(row=r, column=col).value) == target:
                log_debug(f"[search] find_row_by_value found at row={r}, col={col}")
                return r
    log_debug(f"[search] find_row_by_value not found -> None")
    return None


# ----------------------------- UI Components -----------------------------

class BaseCellFrame(ttk.Frame):
    """A professional track configuration frame."""

    def __init__(self, parent, index_changed_cb, remove_cb):
        super().__init__(parent, style="Card.TFrame")
        self.index_changed_cb = index_changed_cb
        self.remove_cb = remove_cb

        self.configure(padding=15)

        # Header with track label and remove button
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", pady=(0, 10))

        track_label = ttk.Label(header_frame, text="Track Configuration",
                                font=("Segoe UI", 10, "bold"))
        track_label.pack(side="left")

        remove_btn = ttk.Button(header_frame, text="Remove Track",
                                command=self._remove_self,
                                style="Danger.TButton")
        remove_btn.pack(side="right")

        # Base cell configuration
        base_frame = ttk.LabelFrame(self, text="Base Cell", padding=10)
        base_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(base_frame, text="Cell Reference:").grid(row=0, column=0, sticky="w", padx=(0, 5))
        self.base_cell_entry = ttk.Entry(base_frame, width=15, font=("Consolas", 9))
        self.base_cell_entry.grid(row=0, column=1, sticky="w")
        self.base_cell_entry.insert(0, "A1")

        # Reference configuration
        ref_frame = ttk.LabelFrame(self, text="Row References (Optional)", padding=10)
        ref_frame.pack(fill="x")

        ref_frame.grid_columnconfigure(1, weight=1)
        ref_frame.grid_columnconfigure(3, weight=1)

        ttk.Label(ref_frame, text="Start Row:").grid(row=0, column=0, sticky="w", padx=(0, 5))
        self.start_ref_entry = ttk.Entry(ref_frame, font=("Consolas", 9))
        self.start_ref_entry.grid(row=0, column=1, sticky="ew", padx=(0, 15))

        ttk.Label(ref_frame, text="End Row:").grid(row=0, column=2, sticky="w", padx=(0, 5))
        self.end_ref_entry = ttk.Entry(ref_frame, font=("Consolas", 9))
        self.end_ref_entry.grid(row=0, column=3, sticky="ew")

        # Help text
        help_frame = ttk.Frame(self)
        help_frame.pack(fill="x", pady=(8, 0))
        help_text = "Tip: Include sheet name like 'Sheet1!C13' for cross-sheet references"
        ttk.Label(help_frame, text=help_text, foreground="#666666",
                  font=("Segoe UI", 8)).pack(anchor="w")

        # Bind change events
        self.base_cell_entry.bind("<FocusOut>", lambda e: self.index_changed_cb())
        self.base_cell_entry.bind("<KeyRelease>", lambda e: self.root.after_idle(self.index_changed_cb))

    def _remove_self(self):
        self.remove_cb(self)

    def get_label(self) -> str:
        val = self.base_cell_entry.get().strip() or "(unset)"
        return val

    def get_data(self) -> Dict[str, str]:
        return {
            "base_cell": self.base_cell_entry.get().strip(),
            "start_ref": self.start_ref_entry.get().strip(),
            "end_ref": self.end_ref_entry.get().strip(),
        }

    def set_data(self, data: Dict[str, str]):
        self.base_cell_entry.delete(0, tk.END)
        self.base_cell_entry.insert(0, data.get("base_cell", "A1"))
        self.start_ref_entry.delete(0, tk.END)
        self.start_ref_entry.insert(0, data.get("start_ref", ""))
        self.end_ref_entry.delete(0, tk.END)
        self.end_ref_entry.insert(0, data.get("end_ref", ""))


class SourceFileFrame(ttk.Frame):
    def __init__(self, parent, on_remove_callback, base_cells_provider):
        super().__init__(parent, style="Card.TFrame")
        self.on_remove_callback = on_remove_callback
        self.base_cells_provider = base_cells_provider
        self.file_path = ""

        self.configure(padding=15)

        # Header with source label and remove button
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", pady=(0, 12))

        source_label = ttk.Label(header_frame, text="Source File Configuration",
            font=("Segoe UI", 10, "bold"))
        source_label.pack(side="left")

        remove_btn = ttk.Button(header_frame, text="Remove Source",
                                command=self.remove_self,
                                style="Danger.TButton")
        remove_btn.pack(side="right")

        # Title configuration
        title_frame = ttk.LabelFrame(self, text="Display Title", padding=10)
        title_frame.pack(fill="x", pady=(0, 10))

        self.title_entry = ttk.Entry(title_frame, font=("Segoe UI", 9))
        self.title_entry.pack(fill="x")

        # File selection
        file_frame = ttk.LabelFrame(self, text="Source File", padding=10)
        file_frame.pack(fill="x", pady=(0, 10))

        file_inner = ttk.Frame(file_frame)
        file_inner.pack(fill="x")
        file_inner.grid_columnconfigure(0, weight=1)

        self.file_label = ttk.Label(file_inner, text="No file selected",
                                    foreground="#999999", anchor="w",
                                    font=("Segoe UI", 9))
        self.file_label.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        browse_btn = ttk.Button(file_inner, text="Browse Files",
                                command=self.browse_file,
                                style="Accent.TButton")
        browse_btn.grid(row=0, column=1)

        # Range and track configuration
        config_frame = ttk.LabelFrame(self, text="Configuration", padding=10)
        config_frame.pack(fill="x")

        config_frame.grid_columnconfigure(1, weight=1)
        config_frame.grid_columnconfigure(3, weight=1)

        ttk.Label(config_frame, text="Source Range:").grid(row=0, column=0, sticky="w", padx=(0, 8))
        self.range_entry = ttk.Entry(config_frame, font=("Consolas", 9))
        self.range_entry.grid(row=0, column=1, sticky="ew", padx=(0, 20))

        ttk.Label(config_frame, text="Target Track:").grid(row=0, column=2, sticky="w", padx=(0, 8))
        self.track_var = tk.StringVar()
        self.track_combo = ttk.Combobox(config_frame, textvariable=self.track_var,
                                        state="readonly", font=("Segoe UI", 9))
        self.track_combo.grid(row=0, column=3, sticky="ew")
        self.refresh_tracks()

        # Help text
        help_frame = ttk.Frame(self)
        help_frame.pack(fill="x", pady=(10, 0))
        help_text = "Examples: A1:C10, Sheet1!B5:D15, A23:B?, A23:?56, A23:??"
        ttk.Label(help_frame, text=help_text, foreground="#666666",
                  font=("Segoe UI", 8)).pack(anchor="w")

    def refresh_tracks(self):
        labels = self.base_cells_provider()
        if not labels:
            labels = ["(no tracks available)"]
        self.track_combo["values"] = labels
        if labels and labels[0] != "(no tracks available)":
            self.track_combo.current(0)

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Source Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.file_path = file_path
            filename = os.path.basename(file_path)
            self.file_label.config(text=filename, foreground="#000000")

    def remove_self(self):
        self.on_remove_callback(self)

    def get_data(self):
        return {
            'title': self.title_entry.get().strip(),
            'file_path': self.file_path,
            'range': self.range_entry.get().strip(),
            'track_label': self.track_var.get(),
        }

    def set_data(self, data: Dict[str, str]):
        self.file_path = data.get("file_path", "")
        self.file_label.config(
            text=(os.path.basename(self.file_path) if self.file_path else "No file selected"),
            foreground=("#000000" if self.file_path else "#999999")
        )

        self.title_entry.delete(0, tk.END)
        self.title_entry.insert(0, data.get("title", ""))

        self.range_entry.delete(0, tk.END)
        self.range_entry.insert(0, data.get("range", ""))

    def is_valid(self):
        return bool(self.file_path and self.range_entry.get().strip())


# ----------------------------- Main Application -----------------------------

class RangePasterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Range Paster Pro")
        self.root.geometry("1200x900")

        self.setup_styles()

        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        self.base_file_path = ""
        self.base_frames: List[BaseCellFrame] = []
        self.source_frames: List[SourceFileFrame] = []

        self.create_widgets()
        set_logger(self.log)

    def setup_styles(self):
        """Configure modern styling for the application."""
        style = ttk.Style()
        style.theme_use('clam')

        colors = {
            'bg': '#f8f9fa',
            'card_bg': '#ffffff',
            'primary': '#0066cc',
            'success': '#28a745',
            'danger': '#dc3545',
            'secondary': '#6c757d',
            'border': '#e9ecef'
        }

        style.configure('Card.TFrame',
                        background=colors['card_bg'],
                        relief='solid',
                        borderwidth=1)

        style.configure('Accent.TButton',
                        background=colors['primary'],
                        foreground='white',
                        focuscolor='none')

        style.configure('Success.TButton',
                        background=colors['success'],
                        foreground='white',
                        focuscolor='none')

        style.configure('Danger.TButton',
                        background=colors['danger'],
                        foreground='white',
                        focuscolor='none')

        style.map('Accent.TButton',
                  background=[('active', '#0052a3')])

        style.map('Success.TButton',
                  background=[('active', '#218838')])

        style.map('Danger.TButton',
                  background=[('active', '#c82333')])

    def create_widgets(self):
        main_container = ttk.Frame(self.root)
        main_container.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        # Adjusted row weights for new side-by-side layout
        main_container.grid_rowconfigure(2, weight=1)  # middle section grows
        main_container.grid_rowconfigure(4, weight=1)  # output log grows
        main_container.grid_columnconfigure(0, weight=1)

        # Application header
        header_frame = ttk.Frame(main_container)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 25))

        title_label = ttk.Label(header_frame, text="Excel Range Paster Pro",
                                font=("Segoe UI", 20, "bold"),
                                foreground="#2c3e50")
        title_label.pack(side="left")

        version_label = ttk.Label(header_frame, text="v2.0",
                                  font=("Segoe UI", 10),
                                  foreground="#7f8c8d")
        version_label.pack(side="left", padx=(10, 0))

        # Base file section
        base_section = ttk.LabelFrame(main_container, text="Base Workbook",
                                      padding=15, style="Card.TFrame")
        base_section.grid(row=1, column=0, sticky="ew", pady=(0, 20))
        base_section.grid_columnconfigure(1, weight=1)

        ttk.Label(base_section, text="Target File:",
                  font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky="w", padx=(0, 10))

        self.base_file_label = ttk.Label(base_section, text="No file selected",
                                         foreground="#999999", anchor="w",
                                         font=("Segoe UI", 9))
        self.base_file_label.grid(row=0, column=1, sticky="ew", padx=(0, 15))

        browse_base_btn = ttk.Button(base_section, text="Select Base File",
                                     command=self.browse_base_file,
                                     style="Accent.TButton")
        browse_base_btn.grid(row=0, column=2)

        # ---------------- Side-by-side middle section ----------------
        middle_section = ttk.Frame(main_container)
        middle_section.grid(row=2, column=0, sticky="nsew", pady=(0, 20))
        middle_section.grid_columnconfigure(0, weight=1)
        middle_section.grid_columnconfigure(1, weight=1)
        middle_section.grid_rowconfigure(0, weight=1)

        # Tracks section (left)
        tracks_section = ttk.LabelFrame(middle_section, text="Base Cell Tracks",
                                        padding=10, style="Card.TFrame")
        tracks_section.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        tracks_section.grid_rowconfigure(0, weight=1)
        tracks_section.grid_columnconfigure(0, weight=1)

        tracks_scroll_frame = ttk.Frame(tracks_section)
        tracks_scroll_frame.grid(row=0, column=0, sticky="nsew")
        tracks_scroll_frame.grid_rowconfigure(0, weight=1)
        tracks_scroll_frame.grid_columnconfigure(0, weight=1)

        self.tracks_canvas = tk.Canvas(tracks_scroll_frame, height=200,
                                       background='#f8f9fa', highlightthickness=0)
        self.tracks_scroll = ttk.Scrollbar(tracks_scroll_frame, orient="vertical",
                                           command=self.tracks_canvas.yview)
        self.tracks_container = ttk.Frame(self.tracks_canvas)

        self.tracks_container.bind("<Configure>",
            lambda e: self.tracks_canvas.configure(scrollregion=self.tracks_canvas.bbox("all")))
        self.tracks_canvas.create_window((0, 0), window=self.tracks_container, anchor="nw")
        self.tracks_canvas.configure(yscrollcommand=self.tracks_scroll.set)

        self.tracks_canvas.grid(row=0, column=0, sticky="nsew")
        self.tracks_scroll.grid(row=0, column=1, sticky="ns")

        add_track_frame = ttk.Frame(tracks_section)
        add_track_frame.grid(row=1, column=0, pady=(10, 0), sticky="w")

        add_track_btn = ttk.Button(add_track_frame, text="+ Add New Track",
                                   command=self.add_base_cell,
                                   style="Success.TButton")
        add_track_btn.pack()

        # Source files section (right)
        sources_section = ttk.LabelFrame(middle_section, text="Source Files",
                                         padding=10, style="Card.TFrame")
        sources_section.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        sources_section.grid_rowconfigure(0, weight=1)
        sources_section.grid_columnconfigure(0, weight=1)

        sources_scroll_frame = ttk.Frame(sources_section)
        sources_scroll_frame.grid(row=0, column=0, sticky="nsew")
        sources_scroll_frame.grid_rowconfigure(0, weight=1)
        sources_scroll_frame.grid_columnconfigure(0, weight=1)

        self.src_canvas = tk.Canvas(sources_scroll_frame, height=250,
                                    background='#f8f9fa', highlightthickness=0)
        self.src_scroll = ttk.Scrollbar(sources_scroll_frame, orient="vertical",
                                        command=self.src_canvas.yview)
        self.src_container = ttk.Frame(self.src_canvas)

        self.src_container.bind("<Configure>",
            lambda e: self.src_canvas.configure(scrollregion=self.src_canvas.bbox("all")))
        self.src_canvas.create_window((0, 0), window=self.src_container, anchor="nw")
        self.src_canvas.configure(yscrollcommand=self.src_scroll.set)

        self.src_canvas.grid(row=0, column=0, sticky="nsew")
        self.src_scroll.grid(row=0, column=1, sticky="ns")

        add_source_frame = ttk.Frame(sources_section)
        add_source_frame.grid(row=1, column=0, pady=(10, 0), sticky="w")

        add_source_btn = ttk.Button(add_source_frame, text="+ Add Source File",
                                    command=self.add_source_file,
                                    style="Success.TButton")
        add_source_btn.pack()

        # ---------------- End side-by-side middle section ----------------

        # Action buttons section
        actions_section = ttk.Frame(main_container)
        actions_section.grid(row=3, column=0, sticky="ew", pady=(0, 20))

        primary_actions = ttk.Frame(actions_section)
        primary_actions.pack(side="left")

        process_btn = ttk.Button(primary_actions, text="Process Ranges",
                                 command=self.process_ranges,
                                 style="Accent.TButton",
                                 width=20)
        process_btn.pack(side="left", padx=(0, 15))

        clear_btn = ttk.Button(primary_actions, text="Clear All",
                               command=self.clear_all,
                               style="Danger.TButton")
        clear_btn.pack(side="left", padx=(0, 30))

        workspace_actions = ttk.Frame(actions_section)
        workspace_actions.pack(side="left")

        save_btn = ttk.Button(workspace_actions, text="Save Workspace",
                              command=self.save_workspace)
        save_btn.pack(side="left", padx=(0, 10))

        load_btn = ttk.Button(workspace_actions, text="Load Workspace",
                              command=self.load_workspace)
        load_btn.pack(side="left")

        # Output log section
        output_section = ttk.LabelFrame(main_container, text="Processing Log",
                                        padding=10, style="Card.TFrame")
        output_section.grid(row=4, column=0, sticky="nsew")
        output_section.grid_rowconfigure(0, weight=1)
        output_section.grid_columnconfigure(0, weight=1)

        self.output_text = scrolledtext.ScrolledText(output_section, height=10,
                                                     font=("Consolas", 9),
                                                     background="#f8f9fa",
                                                     foreground="#2c3e50")
        self.output_text.grid(row=0, column=0, sticky="nsew")

        # Initialize with default content
        self.add_base_cell()
        self.add_source_file()

        self._bind_mousewheel()

    def _bind_mousewheel(self):
        """Enable mouse wheel scrolling on canvas widgets."""

        def _on_mousewheel(event):
            canvas = event.widget
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        self.tracks_canvas.bind("<MouseWheel>", _on_mousewheel)
        self.src_canvas.bind("<MouseWheel>", _on_mousewheel)

    # ----- Track Management -----

    def get_track_labels(self) -> List[str]:
        return [bf.get_label() for bf in self.base_frames]

    def on_track_label_changed(self):
        """Refresh source dropdowns when track labels change."""
        labels = self.get_track_labels()
        for sf in self.source_frames:
            current = sf.track_var.get()
            sf.refresh_tracks()
            if current in sf.track_combo["values"]:
                sf.track_var.set(current)

    def add_base_cell(self):
        """Add a new base cell track."""
        frame = BaseCellFrame(self.tracks_container,
                              index_changed_cb=self.on_track_label_changed,
                              remove_cb=self.remove_base_cell)
        frame.pack(fill="x", pady=(0, 10))
        self.base_frames.append(frame)
        self.root.update_idletasks()
        self.on_track_label_changed()

    def remove_base_cell(self, frame: BaseCellFrame):
        """Remove a base cell track with validation."""
        if len(self.base_frames) <= 1:
            messagebox.showinfo("Cannot Remove",
                                "At least one base cell track is required.",
                                icon="warning")
            return

        result = messagebox.askyesno("Confirm Removal",
                                     "Are you sure you want to remove this track?")
        if result:
            frame.destroy()
            self.base_frames.remove(frame)
            self.root.update_idletasks()
            self.on_track_label_changed()

    # ----- Source File Management -----

    def add_source_file(self):
        """Add a new source file configuration."""
        frame = SourceFileFrame(self.src_container,
                                self.remove_source_file,
                                base_cells_provider=self.get_track_labels)
        frame.pack(fill="x", pady=(0, 10))
        self.source_frames.append(frame)
        self.root.update_idletasks()

    def remove_source_file(self, frame):
        """Remove a source file configuration."""
        if len(self.source_frames) > 1:
            result = messagebox.askyesno("Confirm Removal",
                                         "Are you sure you want to remove this source file?")
            if result:
                frame.destroy()
                self.source_frames.remove(frame)
                self.root.update_idletasks()

    # ----- Logging -----

    def log(self, message):
        """Add a timestamped message to the output log."""
        import datetime
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"

        self.output_text.insert(tk.END, formatted_message + "\n")
        self.output_text.see(tk.END)
        self.root.update()

    # ----- File Operations -----

    def browse_base_file(self):
        """Browse and select the base Excel file."""
        file_path = filedialog.askopenfilename(
            title="Select Base Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.base_file_path = file_path
            filename = os.path.basename(file_path)
            self.base_file_label.config(text=filename, foreground="#000000")
            self.log(f"Base file selected: {filename}")

    def clear_all(self):
        """Clear all inputs and reset to default state."""
        result = messagebox.askyesno("Confirm Clear",
                                     "This will clear all current settings. Continue?",
                                     icon="warning")
        if not result:
            return

        self.base_file_path = ""
        self.base_file_label.config(text="No file selected", foreground="#999999")

        for f in self.base_frames:
            f.destroy()
        self.base_frames.clear()
        self.add_base_cell()

        for f in self.source_frames[1:]:
            f.destroy()
        self.source_frames = self.source_frames[:1]

        if self.source_frames:
            sf = self.source_frames[0]
            sf.file_path = ""
            sf.file_label.config(text="No file selected", foreground="#999999")
            sf.range_entry.delete(0, tk.END)
            sf.title_entry.delete(0, tk.END)
            sf.refresh_tracks()

        self.output_text.delete(1.0, tk.END)
        self.log("All settings cleared and reset to defaults")

    # ----- Validation -----

    def validate_inputs(self) -> bool:
        """Comprehensive input validation with user-friendly error messages."""
        if not self.base_file_path:
            messagebox.showerror("Validation Error",
                                 "Please select a base Excel file.",
                                 icon="error")
            return False

        if not os.path.isfile(self.base_file_path):
            messagebox.showerror("File Error",
                                 f"Base file not found:\n{self.base_file_path}",
                                 icon="error")
            return False

        if not self.base_frames:
            messagebox.showerror("Configuration Error",
                                 "Please add at least one base cell track.",
                                 icon="error")
            return False

        for i, bf in enumerate(self.base_frames, start=1):
            data = bf.get_data()
            if not data["base_cell"]:
                messagebox.showerror("Track Error",
                                     f"Track #{i} is missing a base cell reference.",
                                     icon="error")
                return False

        valid_sources = [f for f in self.source_frames if f.is_valid()]
        if not valid_sources:
            messagebox.showerror("Source Error",
                                 "Please add at least one valid source file with a range.",
                                 icon="error")
            return False

        for i, frame in enumerate(valid_sources, start=1):
            if not os.path.isfile(frame.file_path):
                messagebox.showerror("Source File Error",
                                     f"Source file #{i} not found:\n{frame.file_path}",
                                     icon="error")
                return False

        labels = self.get_track_labels()
        for i, sf in enumerate(self.source_frames, start=1):
            if sf.is_valid() and sf.track_var.get() not in labels:
                messagebox.showerror("Track Assignment Error",
                                     f"Source #{i} has an invalid track selection.",
                                     icon="error")
                return False

        return True

    # ----- Workspace Management -----

    def save_workspace(self):
        """Save current workspace configuration to JSON file."""
        save_path = filedialog.asksaveasfilename(
            title="Save Workspace Configuration",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if not save_path:
            return

        data = {
            "version": "2.0",
            "base_file_path": self.base_file_path,
            "tracks": [bf.get_data() for bf in self.base_frames],
            "sources": [sf.get_data() for sf in self.source_frames],
        }

        try:
            with open(save_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            filename = os.path.basename(save_path)
            self.log(f"Workspace configuration saved: {filename}")
            messagebox.showinfo("Save Successful",
                                f"Workspace saved successfully:\n{filename}")
        except Exception as e:
            error_msg = f"Failed to save workspace: {str(e)}"
            self.log(error_msg)
            messagebox.showerror("Save Error", error_msg)

    def load_workspace(self):
        """Load workspace configuration from JSON file."""
        load_path = filedialog.askopenfilename(
            title="Load Workspace Configuration",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if not load_path:
            return

        try:
            with open(load_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            error_msg = f"Failed to load workspace: {str(e)}"
            messagebox.showerror("Load Error", error_msg)
            return

        for f in self.base_frames:
            f.destroy()
        self.base_frames.clear()

        for f in self.source_frames:
            f.destroy()
        self.source_frames.clear()

        self.base_file_path = data.get("base_file_path", "")
        self.base_file_label.config(
            text=(os.path.basename(self.base_file_path) if self.base_file_path else "No file selected"),
            foreground=("#000000" if self.base_file_path else "#999999")
        )

        tracks = data.get("tracks", [])
        if not tracks:
            self.add_base_cell()
        else:
            for track_data in tracks:
                self.add_base_cell()
                self.base_frames[-1].set_data(track_data)

        sources = data.get("sources", [])
        if not sources:
            self.add_source_file()
        else:
            for source_data in sources:
                self.add_source_file()
                self.source_frames[-1].set_data(source_data)

            labels = self.get_track_labels()
            for sf, sdata in zip(self.source_frames, sources):
                track_label = sdata.get("track_label", "")
                if track_label in labels:
                    sf.track_var.set(track_label)
                elif labels:
                    sf.track_combo.current(0)

        self.root.update_idletasks()
        filename = os.path.basename(load_path)
        self.log(f"Workspace configuration loaded: {filename}")
        messagebox.showinfo("Load Successful",
                            f"Workspace loaded successfully:\n{filename}")

    # ----- Main Processing -----

    def process_ranges(self):
        """Main processing function with enhanced error handling and progress feedback."""
        if not self.validate_inputs():
            return

        self.output_text.delete(1.0, tk.END)
        self.log("=" * 60)
        self.log("STARTING EXCEL RANGE PROCESSING")
        self.log("=" * 60)

        try:
            base_filename = os.path.basename(self.base_file_path)
            self.log(f"Loading base workbook: {base_filename}")
            base_wb = load_workbook(self.base_file_path)
            self.log("Base workbook loaded successfully")

            tracks_info = []
            self.log(f"Configuring {len(self.base_frames)} track(s)...")

            for i, bf in enumerate(self.base_frames, start=1):
                track_data = bf.get_data()
                sheet_name, base_ref = parse_sheet_and_ref(track_data["base_cell"])
                base_ws = get_sheet(base_wb, sheet_name)
                base_col, base_row = parse_cell(base_ref)

                tracks_info.append({
                    "ws": base_ws,
                    "base_col": base_col,
                    "base_row": base_row,
                    "current_col": base_col,
                    "fixed_row": base_row,
                    "start_ref": track_data["start_ref"],
                    "end_ref": track_data["end_ref"],
                })

                self.log(f"  Track {i}: {track_data['base_cell']} "
                         f"(col={base_col}, row={base_row})")

            valid_sources = [sf for sf in self.source_frames if sf.is_valid()]
            self.log(f"Processing {len(valid_sources)} source file(s)...")

            sources = []
            for i, sf in enumerate(valid_sources, start=1):
                src_filename = os.path.basename(sf.file_path)
                self.log(f"  Loading source {i}: {src_filename}")

                try:
                    src_wb = load_workbook(sf.file_path, data_only=True)
                    sources.append({
                        "title": sf.title_entry.get().strip(),
                        "workbook": src_wb,
                        "range": sf.range_entry.get().strip(),
                        "filename": src_filename,
                        "track_label": sf.track_var.get(),
                    })
                    self.log(f"    Source {i} loaded successfully")
                except Exception as e:
                    raise Exception(f"Failed to load source file {i} ({src_filename}): {str(e)}")

            def read_value_from_base_ref(ref_str: str) -> Optional[Any]:
                if not ref_str:
                    return None
                sh, cellref = parse_sheet_and_ref(ref_str)
                ws = get_sheet(base_wb, sh)
                c, r = parse_cell(cellref)
                return ws.cell(row=r, column=c).value

            self.log("-" * 40)
            self.log("PROCESSING SOURCE FILES")
            self.log("-" * 40)

            for i, src in enumerate(sources, start=1):
                labels = self.get_track_labels()
                try:
                    track_idx = labels.index(src["track_label"])
                except ValueError:
                    track_idx = 0

                track = tracks_info[track_idx]
                base_ws = track["ws"]
                current_col = track["current_col"]
                fixed_row = track["fixed_row"]

                title = src.get("title") or ""
                display_name = f"{title} ({src['filename']})" if title else src['filename']

                self.log(f"Processing source {i}: {display_name}")
                self.log(f"  Range: {src['range']}")
                self.log(f"  Target track: {track_idx + 1} [{src['track_label']}]")

                src_sheet_name, src_ref = parse_sheet_and_ref(src["range"])
                src_ws = get_sheet(src["workbook"], src_sheet_name)

                bounds: Tuple[int, int, int, int]
                if ":" in src_ref:
                    left, right = [x.strip() for x in src_ref.split(":", 1)]
                    left_col_tok, left_row_tok = split_col_row(left)
                    end_col_tok, end_row_tok = split_col_row(right)

                    if left_col_tok is None:
                        raise ValueError(f'Invalid start column in "{left}"')
                    if left_col_tok == "?":
                        start_col = 1
                    else:
                        start_col = column_index_from_string(left_col_tok)

                    def ref_row_or_none(ref_key: str) -> Optional[int]:
                        ref_str = track.get(ref_key)
                        if not ref_str:
                            return None
                        val = read_value_from_base_ref(ref_str)
                        if val is None or normalize_value(val) == "":
                            return None
                        return find_row_by_value(src_ws, val)

                    if left_row_tok is None:
                        start_row = 1
                    elif left_row_tok == "?":
                        r = ref_row_or_none("start_ref")
                        if r is not None:
                            start_row = r
                        else:
                            start_row = first_nonempty_row_on_col(src_ws, start_col, 1)
                    else:
                        start_row = int(left_row_tok)

                    if end_col_tok is None:
                        end_col = start_col
                    elif end_col_tok == "?":
                        provisional_min_r = start_row
                        provisional_max_r = src_ws.max_row
                        end_col = last_nonempty_col_in_row_range(src_ws, start_col,
                                                                 provisional_min_r, provisional_max_r)
                    else:
                        end_col = column_index_from_string(end_col_tok)

                    if end_row_tok is None:
                        end_row = start_row
                    elif end_row_tok == "?":
                        r_end = ref_row_or_none("end_ref")
                        if r_end is not None:
                            end_row = r_end
                        else:
                            min_c_tmp, max_c_tmp = (start_col, end_col) if start_col <= end_col else (
                            end_col, start_col)
                            end_row = last_nonempty_row_in_col_range(src_ws, min_c_tmp, max_c_tmp, start_row)
                    else:
                        end_row = int(end_row_tok)

                    if end_col_tok == "?":
                        min_r_final = min(start_row, end_row)
                        max_r_final = max(start_row, end_row)
                        end_col = last_nonempty_col_in_row_range(src_ws, start_col, min_r_final, max_r_final)

                    min_c, max_c = sorted((start_col, end_col))
                    min_r, max_r = sorted((start_row, end_row))
                    bounds = (min_c, min_r, max_c, max_r)

                else:
                    c, r = parse_cell(src_ref)
                    bounds = (c, r, c, r)

                width, height = copy_values(
                    src_ws=src_ws,
                    dst_ws=base_ws,
                    src_bounds=bounds,
                    dst_row=fixed_row,
                    dst_col=current_col
                )

                start_letter = get_column_letter(current_col)
                end_letter = get_column_letter(current_col + width - 1)

                if height <= 1:
                    paste_range = f"{start_letter}{fixed_row}:{end_letter}{fixed_row}"
                else:
                    paste_range = f"{start_letter}{fixed_row}:{end_letter}{fixed_row + height - 1}"

                self.log(
                    f"  Resolved range: {format_range_a1(bounds[0], bounds[1], bounds[2], bounds[3], src_sheet_name)}")
                self.log(f"  Copied {width}x{height} cells to {paste_range}")

                track["current_col"] = current_col + width

            self.log("-" * 40)
            self.log("SAVING RESULTS")
            self.log("-" * 40)
            base_wb.save(self.base_file_path)
            self.log(f"Changes saved to: {base_filename}")

            self.log("=" * 60)
            self.log("PROCESSING COMPLETED SUCCESSFULLY!")
            self.log("=" * 60)

            messagebox.showinfo("Processing Complete",
                                f"Range processing completed successfully!\n\n"
                                f"Results saved to: {base_filename}",
                                icon="info")

        except Exception as e:
            error_msg = f"Error during processing: {str(e)}"
            self.log("=" * 60)
            self.log("PROCESSING FAILED")
            self.log("=" * 60)
            self.log(error_msg)
            messagebox.showerror("Processing Error",
                                 f"An error occurred during processing:\n\n{str(e)}",
                                 icon="error")


def main():
    """Initialize and run the application."""
    root = tk.Tk()

    try:
        root.iconbitmap(default="icon.ico")
    except:
        pass

    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")

    app = RangePasterGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
